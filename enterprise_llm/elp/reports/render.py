"""
Rendering a result set into deliverable formats.

Markdown and CSV are produced directly. PDF goes through the platform's
LaTeX service, which is the report builder already in this codebase - the
same sandbox, the same templates, the same escaping rules.
"""

from __future__ import annotations

import csv
import html
import io
import logging
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

from ..config import ReportSettings, get_settings
from ..latex.render import LatexError, get_renderer
from .datasource import QueryResult

log = logging.getLogger(__name__)

SUPPORTED_FORMATS = ("markdown", "csv", "html", "pdf", "json", "xlsx")

# A PDF wider than this stops being readable in landscape. The existing
# NAMIS generator caps at the same point and tells the reader to use
# Excel or CSV for the full set.
PDF_MAX_COLUMNS = 12

# LaTeX special characters, escaped so a part number containing an
# underscore does not blow up the compile.
_LATEX_ESCAPES = {
    "\\": r"\textbackslash{}",
    "&": r"\&", "%": r"\%", "$": r"\$", "#": r"\#",
    "_": r"\_", "{": r"\{", "}": r"\}",
    "~": r"\textasciitilde{}", "^": r"\textasciicircum{}",
}


@dataclass
class Artifact:
    format: str
    path: str
    bytes: int

    def to_dict(self) -> dict:
        return {"format": self.format, "path": self.path, "bytes": self.bytes}


def _escape_latex(value: str) -> str:
    return "".join(_LATEX_ESCAPES.get(character, character) for character in value)


def _cell(value, max_chars: int) -> str:
    text = "" if value is None else str(value)
    if len(text) > max_chars:
        text = text[: max_chars] + "…"
    return text


def to_markdown(
    result: QueryResult,
    *,
    title: str,
    request_text: str = "",
    narrative: str = "",
    generated_at: datetime | None = None,
    settings: ReportSettings | None = None,
    max_table_rows: int = 500,
) -> str:
    settings = settings or get_settings().reports
    generated_at = generated_at or datetime.now(UTC)

    lines = [f"# {title}", ""]
    lines.append(
        f"*Generated {generated_at.strftime('%Y-%m-%d %H:%M UTC')} · "
        f"{result.row_count} row(s) · {result.duration_ms:.0f} ms*"
    )
    lines.append("")

    if request_text:
        lines += ["**Requested:** " + request_text.strip(), ""]
    if narrative:
        lines += ["## Summary", "", narrative.strip(), ""]

    lines.append("## Results")
    lines.append("")
    if not result.rows:
        lines.append("_No rows returned._")
    else:
        # Escape pipes so a value containing one does not break the table.
        def esc(text: str) -> str:
            return text.replace("|", "\\|").replace("\n", " ")

        lines.append("| " + " | ".join(esc(c) for c in result.columns) + " |")
        lines.append("| " + " | ".join("---" for _ in result.columns) + " |")
        for row in result.rows[:max_table_rows]:
            lines.append(
                "| "
                + " | ".join(esc(_cell(v, settings.max_cell_chars)) for v in row)
                + " |"
            )
        if result.row_count > max_table_rows:
            lines.append("")
            lines.append(
                f"_Showing the first {max_table_rows} of {result.row_count} rows. "
                "The full result is in the CSV artifact._"
            )

    if result.redacted_columns:
        lines += [
            "",
            "> Columns withheld for privacy: "
            + ", ".join(result.redacted_columns),
        ]
    if result.warnings:
        lines += ["", "## Notes", ""] + [f"- {w}" for w in result.warnings]

    return "\n".join(lines) + "\n"


def to_csv(result: QueryResult) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(result.columns)
    for row in result.rows:
        writer.writerow(["" if v is None else v for v in row])
    return buffer.getvalue()


def to_html(
    result: QueryResult,
    *,
    title: str,
    request_text: str = "",
    narrative: str = "",
    generated_at: datetime | None = None,
    settings: ReportSettings | None = None,
) -> str:
    settings = settings or get_settings().reports
    generated_at = generated_at or datetime.now(UTC)
    esc = html.escape

    head = (
        "<style>"
        "body{font:14px/1.55 system-ui,-apple-system,Segoe UI,sans-serif;"
        "margin:2rem auto;max-width:1100px;padding:0 1rem;color:#1a1a1a}"
        "h1{font-size:1.6rem;margin-bottom:.25rem}"
        ".meta{color:#666;font-size:.85rem;margin-bottom:1.5rem}"
        ".summary{background:#f6f8fa;border-left:3px solid #0969da;"
        "padding:.75rem 1rem;margin:1rem 0;white-space:pre-wrap}"
        "table{border-collapse:collapse;width:100%;font-size:.85rem}"
        "th,td{border:1px solid #d0d7de;padding:.35rem .6rem;text-align:left}"
        "th{background:#f6f8fa;position:sticky;top:0}"
        "tr:nth-child(even){background:#fafbfc}"
        ".note{color:#7a5900;background:#fff8c5;padding:.5rem .75rem;"
        "border-radius:4px;margin-top:1rem;font-size:.85rem}"
        "@media(prefers-color-scheme:dark){body{background:#0d1117;color:#e6edf3}"
        "th{background:#161b22}tr:nth-child(even){background:#161b22}"
        "th,td{border-color:#30363d}.summary{background:#161b22}}"
        "</style>"
    )

    parts = [
        head,
        f"<h1>{esc(title)}</h1>",
        f'<div class="meta">Generated {generated_at.strftime("%Y-%m-%d %H:%M UTC")} '
        f"&middot; {result.row_count} row(s) &middot; {result.duration_ms:.0f} ms</div>",
    ]
    if request_text:
        parts.append(f"<p><strong>Requested:</strong> {esc(request_text)}</p>")
    if narrative:
        parts.append(f'<div class="summary">{esc(narrative)}</div>')

    if result.rows:
        parts.append("<table><thead><tr>")
        parts += [f"<th>{esc(c)}</th>" for c in result.columns]
        parts.append("</tr></thead><tbody>")
        for row in result.rows:
            parts.append("<tr>")
            parts += [
                f"<td>{esc(_cell(v, settings.max_cell_chars))}</td>" for v in row
            ]
            parts.append("</tr>")
        parts.append("</tbody></table>")
    else:
        parts.append("<p><em>No rows returned.</em></p>")

    for warning in result.warnings:
        parts.append(f'<div class="note">{esc(warning)}</div>')

    return f"<!doctype html><html><head><meta charset='utf-8'><title>{esc(title)}</title>{''.join(parts[:1])}</head><body>{''.join(parts[1:])}</body></html>"


def to_latex(
    result: QueryResult,
    *,
    title: str,
    request_text: str = "",
    narrative: str = "",
    generated_at: datetime | None = None,
    settings: ReportSettings | None = None,
    max_table_rows: int = 300,
) -> str:
    settings = settings or get_settings().reports
    generated_at = generated_at or datetime.now(UTC)

    visible = min(len(result.columns), PDF_MAX_COLUMNS)
    columns = result.columns[:visible]
    column_spec = "l" * visible if columns else "l"
    header = " & ".join(f"\\textbf{{{_escape_latex(c)}}}" for c in columns)

    body_rows = []
    for row in result.rows[:max_table_rows]:
        cells = [_escape_latex(_cell(v, 80)) for v in row[:visible]]
        body_rows.append(" & ".join(cells) + r" \\")

    narrative_block = ""
    if narrative:
        narrative_block = (
            "\\section*{Summary}\n" + _escape_latex(narrative) + "\n\n"
        )

    truncation_note = ""
    if result.row_count > max_table_rows:
        truncation_note = (
            "\n\\emph{Showing the first " + str(max_table_rows) + " of "
            + str(result.row_count)
            + " rows; the complete result is in the CSV artifact.}\n"
        )

    # Every backslash-bearing fragment is built before the f-string:
    # Python 3.11 forbids backslashes inside f-string expressions.
    requested_block = ""
    if request_text:
        requested_block = r"\textbf{Requested:} " + _escape_latex(request_text) + "\n"

    empty_row = r"\multicolumn{1}{l}{No rows returned.} \\"
    table_body = "\n".join(body_rows) if body_rows else empty_row
    timestamp = generated_at.strftime("%Y-%m-%d %H:%M UTC")
    escaped_title = _escape_latex(title)

    preamble = (
        "\\documentclass[10pt,a4paper]{article}\n"
        "\\usepackage[margin=18mm,landscape]{geometry}\n"
        "\\usepackage{booktabs}\n"
        "\\usepackage{longtable}\n"
        "\\usepackage{array}\n"
        "\\usepackage[hidelinks]{hyperref}\n"
        "\\usepackage{parskip}\n"
    )
    heading = (
        "\\begin{document}\n\n"
        "\\begin{center}\n"
        "  {\\LARGE\\bfseries " + escaped_title + "}\\\\[3pt]\n"
        "  {\\small Generated " + timestamp + " --- "
        + str(result.row_count) + " row(s)}\n"
        "\\end{center}\n\n"
    )
    table = (
        "\\section*{Results}\n"
        "{\\footnotesize\n"
        "\\begin{longtable}{" + column_spec + "}\n"
        "\\toprule\n"
        + header + " \\\\\n"
        "\\midrule\n"
        "\\endhead\n"
        + table_body + "\n"
        "\\bottomrule\n"
        "\\end{longtable}\n"
        "}\n"
    )

    return (
        preamble
        + heading
        + requested_block
        + narrative_block
        + table
        + truncation_note
        + "\n\\end{document}\n"
    )


def _looks_numeric(value) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float, Decimal))


def to_xlsx(
    result: QueryResult,
    path: Path,
    *,
    title: str,
    generated_at: datetime | None = None,
) -> None:
    """
    Write a spreadsheet with real types.

    The point of Excel over CSV is that dates sort as dates and numbers sum
    correctly - a CSV of the same data needs cleaning before anyone can pivot
    it. Values are written in their native type rather than stringified, the
    header is frozen and filterable, and columns are sized to their content.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "openpyxl is required for Excel export; install the 'excel' extra"
        ) from exc

    generated_at = generated_at or datetime.now(UTC)
    workbook = Workbook()
    sheet = workbook.active
    # Excel refuses sheet names over 31 chars or containing []:*?/\
    sheet.title = re.sub(r"[\[\]:*?/\\]", "-", title)[:31] or "Report"

    sheet.append([title])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append([
        f"Generated {generated_at.strftime('%Y-%m-%d %H:%M UTC')} - "
        f"{result.row_count} row(s)"
    ])
    sheet["A2"].font = Font(italic=True, size=9, color="666666")
    sheet.append([])

    header_row = 4
    sheet.append(list(result.columns))
    header_fill = PatternFill("solid", fgColor="DDE5F0")
    for index in range(1, len(result.columns) + 1):
        cell = sheet.cell(row=header_row, column=index)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in result.rows:
        sheet.append([
            # datetime and Decimal round-trip natively; everything else is
            # written as-is so Excel infers the type.
            value if (value is None or _looks_numeric(value)
                      or isinstance(value, (datetime, date))) else str(value)
            for value in row
        ])

    if result.columns:
        span = f"A{header_row}:{get_column_letter(len(result.columns))}{header_row + result.row_count}"
        sheet.auto_filter.ref = span
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

        for index, column in enumerate(result.columns, start=1):
            widest = len(str(column))
            for row in result.rows[:200]:
                if index - 1 < len(row) and row[index - 1] is not None:
                    widest = max(widest, len(str(row[index - 1])))
            sheet.column_dimensions[get_column_letter(index)].width = min(
                50, max(10, widest + 2)
            )

    # Landscape, fit to page width: these reports are wide.
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = f"{header_row}:{header_row}"

    workbook.save(str(path))


async def write_artifacts(
    result: QueryResult,
    *,
    run_id: str,
    title: str,
    formats: list[str],
    request_text: str = "",
    narrative: str = "",
    settings: ReportSettings | None = None,
) -> tuple[list[Artifact], list[str]]:
    """
    Render and persist every requested format.

    A format that fails to render is reported as a warning rather than
    failing the run: a PDF that will not compile should not cost you the CSV
    that would have answered the question.
    """
    settings = settings or get_settings().reports
    directory = Path(settings.artifact_dir) / run_id
    directory.mkdir(parents=True, exist_ok=True)

    generated_at = datetime.now(UTC)
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "-", title).strip("-").lower() or "report"

    artifacts: list[Artifact] = []
    warnings: list[str] = []

    for output_format in formats:
        output_format = output_format.lower().strip()
        if output_format not in SUPPORTED_FORMATS:
            warnings.append(
                f"unsupported output format '{output_format}'; supported: "
                + ", ".join(SUPPORTED_FORMATS)
            )
            continue

        try:
            if output_format == "markdown":
                path = directory / f"{safe_name}.md"
                path.write_text(
                    to_markdown(
                        result, title=title, request_text=request_text,
                        narrative=narrative, generated_at=generated_at, settings=settings,
                    ),
                    encoding="utf-8",
                )
            elif output_format == "csv":
                path = directory / f"{safe_name}.csv"
                path.write_text(to_csv(result), encoding="utf-8")
            elif output_format == "html":
                path = directory / f"{safe_name}.html"
                path.write_text(
                    to_html(
                        result, title=title, request_text=request_text,
                        narrative=narrative, generated_at=generated_at, settings=settings,
                    ),
                    encoding="utf-8",
                )
            elif output_format == "json":
                import json

                path = directory / f"{safe_name}.json"
                path.write_text(
                    json.dumps(
                        {
                            "title": title,
                            "generated_at": generated_at.isoformat(),
                            "request": request_text,
                            "narrative": narrative,
                            "columns": result.columns,
                            "rows": result.rows,
                            "row_count": result.row_count,
                            "warnings": result.warnings,
                        },
                        indent=2,
                        default=str,
                    ),
                    encoding="utf-8",
                )
            elif output_format == "xlsx":
                path = directory / f"{safe_name}.xlsx"
                to_xlsx(result, path, title=title, generated_at=generated_at)
            else:  # pdf
                if len(result.columns) > PDF_MAX_COLUMNS:
                    warnings.append(
                        f"the PDF shows the first {PDF_MAX_COLUMNS} of "
                        f"{len(result.columns)} columns so it stays readable; "
                        "use the Excel or CSV artifact for the full set"
                    )
                source = to_latex(
                    result, title=title, request_text=request_text,
                    narrative=narrative, generated_at=generated_at, settings=settings,
                )
                compiled = await get_renderer().compile(source)
                if not compiled.success:
                    warnings.append(
                        "the PDF failed to compile: "
                        + "; ".join(compiled.errors[:2] or ["unknown LaTeX error"])
                    )
                    continue
                path = directory / f"{safe_name}.pdf"
                path.write_bytes(Path(compiled.pdf_path).read_bytes())

            artifacts.append(
                Artifact(
                    format=output_format,
                    path=str(path),
                    bytes=path.stat().st_size,
                )
            )
        except LatexError as exc:
            warnings.append(f"PDF rendering failed: {exc}")
        except OSError as exc:
            warnings.append(f"could not write the {output_format} artifact: {exc}")

    return artifacts, warnings
