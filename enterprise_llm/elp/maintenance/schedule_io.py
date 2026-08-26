"""
Importing the standard maintenance schedule.

The customer supplies their maintenance programme as a spreadsheet or CSV
export.  Column names vary between operators, so the importer matches on a
set of aliases rather than demanding one fixed layout, validates every row,
and reports what it rejected instead of silently dropping it - a task card
that fails to import is a limit nobody is tracking.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import MaintenanceTask

log = logging.getLogger(__name__)

# Accepted spellings for each field, lower-cased and stripped of separators.
_ALIASES: dict[str, tuple[str, ...]] = {
    "task_code": ("taskcode", "task", "taskno", "tasknumber", "cardno", "card", "reference", "ref"),
    "title": ("title", "taskdescription", "description", "name", "tasktitle"),
    "description": ("detail", "details", "longdescription", "notes", "remarks"),
    "ata_chapter": ("ata", "atachapter", "chapter", "atacode"),
    "task_type": ("tasktype", "type", "category", "maintenancetype"),
    "interval_flight_hours": ("intervalfh", "fh", "flighthours", "hours", "hrs", "intervalhours", "fhinterval"),
    "interval_cycles": ("intervalcycles", "cycles", "cycleinterval", "cyc"),
    "interval_landings": ("intervallandings", "landings", "ldg", "landinginterval"),
    "interval_calendar_days": ("intervaldays", "days", "calendardays", "calendar", "dayinterval"),
    "interval_calendar_months": ("months", "intervalmonths", "calendarmonths", "monthinterval", "mo"),
    "tolerance_flight_hours": ("tolerancefh", "tolerancehours", "fhtolerance"),
    "tolerance_calendar_days": ("tolerancedays", "daytolerance", "tolerance"),
    "estimated_man_hours": ("manhours", "mhrs", "labourhours", "laborhours", "estimatedmanhours"),
    "estimated_downtime_hours": ("downtimehours", "downtime", "elapsedhours", "groundtime"),
    "technicians_required": ("technicians", "techs", "manpower", "crewsize"),
    "requires_hangar": ("hangar", "requireshangar", "hangarrequired"),
    "is_airworthiness_limitation": ("awl", "airworthinesslimitation", "lifelimited", "limitation"),
    "can_be_deferred": ("candefer", "deferrable", "canbedeferred"),
    "max_deferral_days": ("maxdeferraldays", "deferrallimit", "maxdeferral"),
    "applicable_models": ("model", "models", "aircraftmodel", "applicablemodels"),
    "applicable_configurations": ("configuration", "configurations", "config", "effectivity"),
    "applicable_serials": ("serial", "serials", "serialnumbers", "sn"),
    "source_document_key": ("sourcedocument", "document", "manual", "sourcedoc", "docref"),
    "source_reference": ("sourcereference", "section", "clause", "paragraph", "reference2"),
}

_TRUE = {"y", "yes", "true", "1", "x", "required", "mandatory"}
_FALSE = {"n", "no", "false", "0", "", "-", "optional"}


def _normalise(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def _build_column_map(headers: list[str]) -> dict[str, str]:
    """Map each source column onto a task field."""
    mapping: dict[str, str] = {}
    for header in headers:
        key = _normalise(header)
        if not key:
            continue
        for field_name, aliases in _ALIASES.items():
            if key == _normalise(field_name) or key in aliases:
                mapping[header] = field_name
                break
    return mapping


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "n/a", "na"}:
        return None
    match = re.search(r"-?\d+(\.\d+)?", text)
    return float(match.group()) if match else None


def _to_int(value: Any) -> int | None:
    result = _to_float(value)
    return int(result) if result is not None else None


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in _TRUE:
        return True
    if text in _FALSE:
        return False
    return default


def _to_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if not text or text.lower() in {"all", "-", "n/a"}:
        return []
    return [part.strip() for part in re.split(r"[;,/|]", text) if part.strip()]


@dataclass
class ImportIssue:
    row: int
    task_code: str
    problem: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    issues: list[ImportIssue] = field(default_factory=list)
    unmapped_columns: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "unmapped_columns": self.unmapped_columns,
            "issues": [
                {"row": i.row, "task_code": i.task_code, "problem": i.problem}
                for i in self.issues
            ],
        }


def read_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Read CSV, JSON or XLSX into a list of raw row dictionaries."""
    suffix = path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            rows = [dict(row) for row in reader]
            return rows, list(reader.fieldnames or [])

    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        rows = data.get("tasks", data) if isinstance(data, dict) else data
        if not isinstance(rows, list):
            raise ValueError("JSON schedule must be a list of task objects")
        headers = sorted({key for row in rows for key in row})
        return rows, headers

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError(
                "openpyxl is required to read .xlsx schedules; "
                "export to CSV or install it"
            ) from exc
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        sheet = workbook.active
        raw = list(sheet.iter_rows(values_only=True))
        if not raw:
            return [], []
        headers = [str(h) if h is not None else "" for h in raw[0]]
        rows = [
            {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            for row in raw[1:]
            if any(cell is not None for cell in row)
        ]
        return rows, headers

    raise ValueError(
        f"unsupported schedule format '{suffix}'; use .csv, .tsv, .json or .xlsx"
    )


def parse_row(raw: dict[str, Any], column_map: dict[str, str]) -> dict[str, Any]:
    """Translate one source row into task-card fields."""
    values: dict[str, Any] = {}
    for column, field_name in column_map.items():
        if column in raw:
            values[field_name] = raw[column]

    months = _to_int(values.get("interval_calendar_months"))
    calendar_days = _to_int(values.get("interval_calendar_days"))
    if calendar_days is None and months is not None:
        # Calendar intervals in maintenance programmes are months; 30-day
        # months are the usual convention for planning purposes.
        calendar_days = months * 30

    return {
        "task_code": str(values.get("task_code", "")).strip(),
        "title": str(values.get("title", "")).strip(),
        "description": str(values.get("description", "") or "").strip(),
        "ata_chapter": str(values.get("ata_chapter", "") or "").strip(),
        "task_type": str(values.get("task_type", "inspection") or "inspection").strip().lower(),
        "interval_flight_hours": _to_float(values.get("interval_flight_hours")),
        "interval_cycles": _to_int(values.get("interval_cycles")),
        "interval_landings": _to_int(values.get("interval_landings")),
        "interval_calendar_days": calendar_days,
        "tolerance_flight_hours": _to_float(values.get("tolerance_flight_hours")) or 0.0,
        "tolerance_calendar_days": _to_int(values.get("tolerance_calendar_days")) or 0,
        "estimated_man_hours": _to_float(values.get("estimated_man_hours")) or 1.0,
        "estimated_downtime_hours": _to_float(values.get("estimated_downtime_hours")) or 1.0,
        "technicians_required": _to_int(values.get("technicians_required")) or 1,
        "requires_hangar": _to_bool(values.get("requires_hangar")),
        "is_airworthiness_limitation": _to_bool(values.get("is_airworthiness_limitation")),
        "can_be_deferred": _to_bool(values.get("can_be_deferred"), default=True),
        "max_deferral_days": _to_int(values.get("max_deferral_days")) or 0,
        "applicable_models": _to_list(values.get("applicable_models")),
        "applicable_configurations": _to_list(values.get("applicable_configurations")),
        "applicable_serials": _to_list(values.get("applicable_serials")),
        "source_document_key": str(values.get("source_document_key", "") or "").strip(),
        "source_reference": str(values.get("source_reference", "") or "").strip(),
    }


def validate(parsed: dict[str, Any]) -> str | None:
    """Return a problem description, or ``None`` when the row is usable."""
    if not parsed["task_code"]:
        return "no task code"
    if not parsed["title"]:
        return "no title or description"
    has_interval = any(
        parsed[key]
        for key in (
            "interval_flight_hours",
            "interval_cycles",
            "interval_landings",
            "interval_calendar_days",
        )
    )
    if not has_interval and parsed["task_type"] not in {"ad", "sb", "one-time", "onetime"}:
        return (
            "no interval on any basis (flight hours, cycles, landings or calendar) - "
            "a recurring task needs at least one"
        )
    if parsed["is_airworthiness_limitation"] and parsed["can_be_deferred"]:
        # Not fatal, but almost always a data-entry error worth surfacing.
        return None
    return None


async def import_schedule(
    session: AsyncSession,
    path: Path,
    *,
    default_models: list[str] | None = None,
    source_document_key: str = "",
    replace_existing: bool = False,
    dry_run: bool = False,
) -> ImportResult:
    """Load a maintenance programme file into the task-card table."""
    rows, headers = read_rows(path)
    if not rows:
        raise ValueError(f"{path.name} contains no data rows")

    column_map = _build_column_map(headers)
    if "task_code" not in column_map.values():
        raise ValueError(
            "could not find a task code column. Expected one of: "
            + ", ".join(_ALIASES["task_code"])
        )

    result = ImportResult(
        unmapped_columns=[h for h in headers if h and h not in column_map]
    )

    existing_rows = (await session.execute(select(MaintenanceTask))).scalars().all()
    existing = {row.task_code: row for row in existing_rows}
    seen: set[str] = set()

    for index, raw in enumerate(rows, start=2):  # row 1 is the header
        parsed = parse_row(raw, column_map)
        problem = validate(parsed)
        if problem:
            result.skipped += 1
            result.issues.append(
                ImportIssue(row=index, task_code=parsed["task_code"], problem=problem)
            )
            continue

        if parsed["task_code"] in seen:
            result.skipped += 1
            result.issues.append(
                ImportIssue(
                    row=index,
                    task_code=parsed["task_code"],
                    problem="duplicate task code in the source file",
                )
            )
            continue
        seen.add(parsed["task_code"])

        if not parsed["applicable_models"] and default_models:
            parsed["applicable_models"] = list(default_models)
        if not parsed["source_document_key"] and source_document_key:
            parsed["source_document_key"] = source_document_key

        if dry_run:
            result.created += parsed["task_code"] not in existing
            result.updated += parsed["task_code"] in existing
            continue

        row = existing.get(parsed["task_code"])
        if row is None:
            row = MaintenanceTask(task_code=parsed["task_code"])
            session.add(row)
            result.created += 1
        else:
            result.updated += 1

        for key, value in parsed.items():
            if key != "task_code":
                setattr(row, key, value)
        row.active = True

    if replace_existing and not dry_run:
        for task_code, row in existing.items():
            if task_code not in seen:
                # Retire rather than delete: compliance history references it.
                row.active = False

    if not dry_run:
        await session.flush()

    log.info(
        "imported %s: %d created, %d updated, %d skipped",
        path.name, result.created, result.updated, result.skipped,
    )
    return result
